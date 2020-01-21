import requests
import csv
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
import os
from pathlib import Path

class SPI_Data_Loader_Service():
    view_service = None

    def __init__(self, _view_service):
        self.view_service = _view_service

    def update_spi_file(self):
        fetched_data = self.fetch_spi_matches()
        processed_data = self.process_fetched_data(fetched_data)
        self.write_to_xlsx_file(processed_data)
    
    def fetch_spi_matches(self):
        url = "https://projects.fivethirtyeight.com/soccer-api/club/spi_matches.csv"
        response = requests.get(url)
        return response.text

    def process_fetched_data(self, fetched_data):
        first = fetched_data.split('\n')
        corr_array = []
        for row in first:
            col = row.split(',')
            corr_array.append(col)
        return corr_array

    def write_to_xlsx_file(self, processed_data):
        self.view_service.print_progress_bar(0, 100, prefix = 'Load-Data:', suffix = 'Complete', length = 50)
        path = Path(__file__).parent / "../../spi_data/automated_spi.xlsx"
        workbook = load_workbook(filename = path)
        active_workbook = workbook.active
        for r, row in enumerate(processed_data):
            for c, value in enumerate(row):
                progress = int((r / len(processed_data))* 100) 
                self.view_service.print_progress_bar(progress, 100, prefix = 'Load-Data:', suffix = 'Complete', length = 50)
                active_workbook.cell(row = r+1, column = c+1).value = value
        workbook.save(path)



