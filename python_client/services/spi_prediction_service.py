from openpyxl import load_workbook
import datetime
from datetime import timedelta
from enum import Enum
from pathlib import Path

class Filter_Criteria(Enum):
    game_with_result = 1
    game_without_result = 2


class SPI_Prediction_Service:
    view_service = None

    def __init__(self, _view_service):
        self.view_service = _view_service

    def get_ratio_of_ended_games(self, percentage_begin, percentage_end):
        correct_games = 0
        games_in_percentage_range = 0
        games = self.get_games_of_percentage_range(
            percentage_begin, percentage_end)
        filtered_games = self.filter_games(
            Filter_Criteria.game_with_result, games)
        games_in_percentage_range = len(filtered_games)

        for game in filtered_games:
            if (float(game[14]) > float(game[15])) and (game[7] > game[8]):
                correct_games = correct_games + 1
            if (float(game[15]) > float(game[14])) and (game[8] > game[7]):
                correct_games = correct_games + 1
        return correct_games/games_in_percentage_range

    def get_games_of_percentage_range(self, percentage_begin, percentage_end):
        path = Path(__file__).parent / "../../spi_data/automated_spi.xlsx"
        games = []
        workbook = load_workbook(
            filename=path)
        active_workbook = workbook.active

        for row in active_workbook.iter_rows(min_row=2, values_only=True):
            if row[7] != None and row[8] != None:
                if float(row[7]) >= percentage_begin and float(row[7]) <= percentage_end:
                    games.append(row)
                if float(row[8]) >= percentage_begin and float(row[8]) <= percentage_end:
                    games.append(row)
        return games

    def filter_games(self, filter_criteria, games):
        filtered_games = []
        if filter_criteria == Filter_Criteria.game_with_result:
            for game in games:
                if ((game[7] != None) and (game[8] != None) and (game[14] != None) and (game[15] != None)):
                    filtered_games.append(game)

        if filter_criteria == Filter_Criteria.game_without_result:
            for game in games:
                if ((game[7] != None) and (game[8] != None) and (game[14] == None) and (game[15] == None)):
                    filtered_games.append(game)
        return filtered_games

    def get_all_not_started_games_in_percentage_range(self, percentage_begin, percentage_end):
        games = self.get_games_of_percentage_range(
            percentage_begin, percentage_end)
        filtered_games = self.filter_games(
            Filter_Criteria.game_without_result, games)
        return filtered_games

    def eliminate_games_more_than_week(self, games):
        filtered_games = []
        for game in games:
            date = datetime.datetime.strptime(game[0], '%Y-%m-%d')
            if date > datetime.datetime.now() and date < datetime.datetime.now() + timedelta(days=7):
                filtered_games.append(game)
        return filtered_games

    def get_games_to_bet_this_week(self):
        games_to_bet = []
        self.view_service.print_progress_bar(0, 6, prefix = 'Prepare-Data:', suffix = 'Complete', length = 50)
        for i in range(6):
            min_probability = 0.65 + i * 0.05
            max_probability = 0.7 + i * 0.05

            win_lose_ratio = self.get_ratio_of_ended_games(
                min_probability, max_probability)
            min_quote = 1 / win_lose_ratio

            games = self.get_all_not_started_games_in_percentage_range(
                min_probability, max_probability)
            filtered_games = self.eliminate_games_more_than_week(games)

            for game in filtered_games:
                computed_game = self.compute_game_dictionary(game, min_quote)
                games_to_bet.append(computed_game)
            self.view_service.print_progress_bar(i+1, 6, prefix = 'Prepare-Data:', suffix = 'Complete', length = 50)
        
        return games_to_bet

    def compute_game_dictionary(self, game, min_quote):
        game_dict = {}
        winning_side = 'home' if float(
            game[7]) > float(game[8]) else 'away'
        winning_side_probability = game[7] if float(
            game[7]) > float(game[8]) else game[8]
        game_dict['league'] = game[2]
        game_dict['home'] = game[3]
        game_dict['away'] = game[4]
        game_dict['date'] = game[0]
        game_dict['min_quote'] = min_quote
        game_dict['winning_side'] = winning_side
        game_dict['winning_side_probability'] = winning_side_probability
        return game_dict

    def compute_game_information_dictionary(self, game, fixture, best_bookmaker):
        game_information = {}
        game_information['date'] = game['date']
        game_information['home'] = game['home']
        game_information['away'] = game['away']
        game_information['compared_home'] = fixture['homeTeam']['team_name']
        game_information['compared_away'] = fixture['awayTeam']['team_name']
        game_information['minimal_betting_odd'] = game['min_quote']
        game_information['winning_side'] = game['winning_side']
        game_information['winning_side_probability'] = game['winning_side_probability']
        game_information['best_bookmaker'] = best_bookmaker['bookmaker_name']
        game_information['best_bookmaker_odd'] = best_bookmaker['odd']
        return game_information

    def calculate_expectation_value(self, bettable_games):
        expectation_value = 0
        excluded_array = self.exclude_not_working_games()
        for game in bettable_games:
            current_expectation_value = (float(game['winning_side_probability']) * 1) + (1- float(game['winning_side_probability']) * -1)
            expectation_value = expectation_value + current_expectation_value
        return expectation_value

    def exclude_not_working_games(self, bettable_games):
        excluded_array = []
        for game in bettable_games:
            if filtered_game['best_bookmaker'] != '':
                excluded_array.append(game)
        return excluded_array



    